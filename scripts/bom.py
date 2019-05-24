from configReader import *
from genericframe import *
from partdb import *
from digikey import *
from openpyxl import load_workbook


class UpdateBOMApplication(GenericFrame):
    FAVICON = "../assets/favicon.ico"

    def __init__(self, parent=None):
        Frame.__init__(self, parent)
        self.parent = parent
        
        self.parent.title("Partlocater - BOM Updater")
        self.parent.iconbitmap(self.FAVICON)
        self.prod_quantity = StringVar()
        self.report_file = None
        self.updateBOMFrame = LabelFrame(self, text="Update Bill of Materials")
        self.updateBOMFrame.pack(side=TOP, fill=BOTH, expand=YES, pady=4, padx=4)

        self.fileFrame = Frame(self.updateBOMFrame)
        self.fileFrame.pack(side=TOP, fill=X, expand=YES)
        self.filename = StringVar()
        self.fileEntry = Entry(self.fileFrame, textvariable=self.filename, width=60)
        self.fileEntry.pack(side=LEFT, fill=X, expand=YES, pady=4, padx=4)
        self.fileButton = ttk.Button(self.fileFrame, text="...", command=self.open_file, state=NORMAL, width=3)
        self.fileButton.pack(side=LEFT, anchor=W, fill=X, expand=NO)

        val = IntVar()
        self.updateDBCheckbutton = Checkbutton(self.updateBOMFrame, anchor=W,
                                               text="Update Price/Quantity from Digi-Key", variable=val)
        self.updateDBCheckbutton.var = val
        self.updateDBCheckbutton.pack(side=TOP,  fill=X, expand=YES)
        val = IntVar()
        self.bestPriceCheckbutton = Checkbutton(self.updateBOMFrame, anchor=W,
                                                text="Adjust Extended Quantity to Best Price", variable=val)
        self.bestPriceCheckbutton.var = val
        self.bestPriceCheckbutton.pack(side=TOP,  fill=X, expand=YES)
        self.buildFrame = Frame(self.updateBOMFrame)
        self.buildFrame.pack(side=TOP, fill=X, expand=YES)
        val = StringVar()
        self.build_entry = Entry(self.buildFrame, justify=RIGHT, textvariable=val, width=8)
        self.build_entry.var = val
        self.build_entry.pack(side=LEFT, fill=X, expand=NO, pady=4, padx=4)
        self.buildLabel = Label(self.buildFrame, anchor=W, text="Production Quantity")
        self.buildLabel.pack(side=LEFT, fill=X, expand=NO, pady=4)
        self.buttonFrame = Frame(self.updateBOMFrame)
        self.buttonFrame.pack(side=TOP, fill=X, expand=NO, pady=4, padx=4)
        self.clearButton = ttk.Button(self.buttonFrame, text="Clear", state=NORMAL, command=self.do_clear, width=8)
        self.clearButton.pack(side=RIGHT, anchor=W, fill=X, expand=NO)
        self.updateButton = ttk.Button(self.buttonFrame, text="Update", state=NORMAL, command=self.do_update, width=8)
        self.updateButton.pack(side=RIGHT, anchor=W, fill=X, expand=NO)
        self.statusFrame = LabelFrame(self, text="Status")
        self.statusFrame.pack(side=BOTTOM, fill=X, expand=YES, pady=4, padx=6)
        self.status = self.StatusBar(self.statusFrame, self)
        self.status.set("Open Bill of Materials Spreadsheet (.xlsx)")
        self.pack(side=LEFT, fill=BOTH, expand=YES)
        self.do_clear()

        self.BOM_wb = None
        self.BOM_sheet = None
        self.spn_rowcol = None
        self.supplier_rowcol = None
        self.unitprice_rowcol = None
        self.stock_rowcol = None
        self.quantity_rowcol = None
        self.extqty_rowcol = None

    def open_file(self):
        self.status.set("Open BOM File to modify")
        fname = filedialog.askopenfilename(initialdir=".", title="Open .xlsx BOM file",
                                           filetypes=(("Excel", "*.xlsx"), ("all files", "*.*")))
        if not fname:
            self.status.set("")
            return None
        self.filename.set(fname)
        self.report_file = os.path.splitext(fname)[0] + '.txt'
        self.status.set("Opened %s", fname)
        self.parent.focus_force()

    def do_clear(self):
        self.updateDBCheckbutton.var.set(0)
        if 'BestPrice' in Config().bom:
            self.bestPriceCheckbutton.var.set(Config().bom['BestPrice'])
        if 'ProductionQuantity' in Config().bom:
            self.build_entry.config(state=DISABLED)
        self.filename.set("")

    def valid_cell(self, string):
        pattern = re.compile('^[A-Z][0-9]+$')
        return True if pattern.findall(string) else False
        
    def valid_number(self, string):
        pattern = re.compile('^[\\+\\-]?[\\d]+$')
        return True if pattern.findall(string) else False
        
    def search(self, string):
        for i in range(1, self.BOM_sheet.max_row):
            for j in range(1, self.BOM_sheet.max_column):
                if self.BOM_sheet.cell(row=i, column=j).value == string:
                    return i, j
        return None

    def build_part_list(self, supplier_name):
        part_list = []
        row_list = dict()
        for i in range(self.supplier_rowcol[0]+1, self.BOM_sheet.max_row):
            if self.BOM_sheet.cell(row=i, column=self.supplier_rowcol[1]).value == supplier_name:
                part = []
                part.append(self.BOM_sheet.cell(row=i, column=self.spn_rowcol[1]).value)
                part.append(self.BOM_sheet.cell(row=i, column=self.quantity_rowcol[1]).value)
                part_list.append(part)
                row_list[part[0]] = i
        return part_list, row_list

    def report_write(self, format, *args):
        if self.report_file is not None:
            with open(self.report_file, 'a') as f:
                f.write(str(format) % args + "\n")

    def compute_price(self, quantity, price_str, best_price):
        price_list = []
        # convert string to list
        for item in price_str.split(','):
            pair = item.split('=')
            price_list.append([int(pair[0]), float(pair[1])])
        # do first break 
        last_price = price_list[0][1]
        last_break = price_list[0][0]
        if quantity <= last_break:
            return last_break if best_price else quantity, last_price, 0
        # do remaining break
        for i in range(1,len(price_list)):
            price_break = price_list[i][0]
            price = price_list[i][1]
            if price_break > quantity:
                if best_price and ((last_price * quantity) > (price_break * price)):
                    return price_break, price, 0
                else:
                    return quantity, last_price, 0
            last_price = price
        # if it exceeds break
        return quantity, price_list[-1][1], 1

    def do_update(self):
        filename = self.filename.get()
        if not filename:
            self.status.seterror("Requires Bill of Materials File to update")
            return
        try:
            self.BOM_wb = load_workbook(filename=filename)
        except Exception as e:
            self.status.seterror(e)
            return
        if 'ActiveSheet' in Config().bom:
            if Config().bom['ActiveSheet'] not in self.BOM_wb:
                self.status.seterror("BOM doesn't have sheet named %s"%Config().bom['ActiveSheet'])
                return
            self.BOM_sheet = self.BOM_wb[Config().bom['ActiveSheet']]
        else:
            self.BOM_sheet = self.BOM_wb.active
        # we get the build quantity either in the entry box 
        # or from the spreadsheet and put that in the entry box
        if 'ProductionQuantity' in Config().bom:
            pq_cell = Config().bom['ProductionQuantity']
            if not self.valid_cell(pq_cell):
                self.status.seterror("ProductionQuantity value is invalid cell reference")
                return
            pq = self.BOM_sheet[pq_cell].value
            if self.valid_number(pq):
                self.prod_quantity = self.BOM_sheet[pq_cell].value
                self.build_entry.var.set(self.prod_quantity)
            else:
                self.status.seterror("Invalid Production Quantity number in Spreadsheet")
                return
        else:
            pq = self.build_entry.var.get()
            if self.valid_number(pq):
                self.prod_quantity = pq
            else:
                self.status.seterror("Production Quantity is invalid")
                return
        self.spn_rowcol = self.search('Supplier Part Number 1')
        if self.spn_rowcol is None:
            self.status.seterror("Can't find Supplier Part Number Column in Spreadsheet")
            return
        self.supplier_rowcol = self.search('Supplier 1')
        if self.supplier_rowcol is None:
            self.status.seterror("Can't find Supplier Column in Spreadsheet")
            return
        self.unitprice_rowcol = self.search('Supplier Unit Price 1')
        if self.unitprice_rowcol is None:
            self.status.seterror("Can't find Supplier Unit Price 1 Column in Spreadsheet")
            return
            
        self.stock_rowcol = self.search('Supplier Stock 1')
 
        self.quantity_rowcol = self.search('Quantity')
        if self.quantity_rowcol is None:
            self.status.seterror("Can't find Quantity Column in Spreadsheet")
            return
        self.extqty_rowcol = self.search('Extended Quantity')
        if self.extqty_rowcol is None:
            self.status.seterror("Can't find Extended Quantity in Spreadsheet")
            return
        supplier = 'Digi-Key'
        partlist, rowlist = self.build_part_list(supplier)
        self.status.set("Partlist extracted from Spreadsheet")
        try:
            os.remove(self.report_file)
        except:
            pass
        self.report_write("BOM Updater")
        self.report_write("Supplier: %s", supplier)
        if 'ProjectNumber' in Config().bom:
            self.report_write("Project: %s", self.BOM_sheet[Config().bom['ProjectNumber']].value)
        if 'RevisionNumber' in Config().bom:
            self.report_write("Revision: %s", self.BOM_sheet[Config().bom['RevisionNumber']].value)
        self.report_write("%s Line Items: %d", supplier, len(partlist))
        self.report_write("Build Quantity: %d", int(self.prod_quantity))
        self.report_write("")
        if self.updateDBCheckbutton.var.get():
            for part in partlist:
                try:
                    part_json = digikey_get_part(part[0])
                except Exception as e:
                    self.report_write("Digi-Key part number %s is invalid", part[0])
                    part_json = None
                    pass
                try:
                    parameters, category, alt_package, hidden = translate_part_json(part_json)
                except Exception as e:
                    self.handleError("Translate Digi-Key Response Error", e)
                    return
                if parameters is not None:
                    Config().loaded_db.update_priceinfo(category, parameters['Pricing 1'], parameters['Supplier Stock 1'], parameters['Supplier Part Number 1'])
                self.status.set("%s updated in database", parameters['Supplier Part Number 1'])

        price_list = Config().loaded_db.get_priceinfo(Config().tables, Config().loaded_metadb.name, supplier, partlist)
        
        best_price = self.bestPriceCheckbutton.var.get()
        for part in price_list:
            row = rowlist[part['Supplier Part Number 1']]
            part_quantity = self.BOM_sheet.cell(row=row, column=self.quantity_rowcol[1]).value
            extended_quantity, price, warn_flag = self.compute_price(int(part_quantity)*int(self.prod_quantity),
                                                                     part['Pricing 1'], best_price)
            if warn_flag:
                self.report_write("Recommend Volume Packaging for Part %s Row %d", part['Supplier Part Number 1'], row)
            self.BOM_sheet.cell(row=row, column=self.unitprice_rowcol[1]).value = price
            self.BOM_sheet.cell(row=row, column=self.extqty_rowcol[1]).value = extended_quantity
            if 'Supplier Stock 1' in part and part['Supplier Stock 1'] is not None:
                stock = int(part['Supplier Stock 1'])
                if self.stock_rowcol is not None:
                    self.BOM_sheet.cell(row=row, column=self.stock_rowcol[1]).value = stock
                    self.report_write("Stock for %s is %s (%s)", part['Supplier Part Number 1'],
                                      str(stock), type(stock))
                if stock == 0:
                    self.report_write("Out of stock for Part %s Row %d", part['Supplier Part Number 1'], row)
        try:
            self.BOM_wb.save(self.filename.get())
        except Exception as e:
            self.status.seterror(e)
            return
        self.status.set("BOM Updated")
        self.updateDBCheckbutton.var.set(0)
